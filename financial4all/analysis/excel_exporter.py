# financial4all/analysis/excel_exporter.py
"""
Excel export functionality for financial analysis reports.

This module provides functionality for exporting comprehensive financial
analysis reports to Excel with professional formatting.
"""

import pandas as pd
from typing import Union, Optional
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from financial4all.analysis.report_generator import FinancialAnalysisReport
from financial4all.core import log

if not OPENPYXL_AVAILABLE:
    log.warning("openpyxl not available. Excel export will not work.")


class ExcelExporter:
    """
    Exports financial analysis reports to Excel.

    This class handles formatting, multi-sheet workbooks, and professional
    presentation of financial data.
    """

    def export_analysis(
        self,
        report: FinancialAnalysisReport,
        file_path_or_buffer: Union[str, BytesIO],
        include_charts: bool = False,
    ) -> None:
        """
        Export analysis to Excel with formatting.

        Args:
            report: FinancialAnalysisReport object
            file_path_or_buffer: File path (string) or BytesIO buffer
            include_charts: Whether to include charts (default: False)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. Install it with: pip install openpyxl"
            )

        # Generate all report data
        report_data = report.generate_report()

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # 1. Summary Sheet
        self._create_summary_sheet(wb, report, report_data)

        # 2. Income Statement Sheet
        multi_year = report_data.get("multi_year_comparison", {})
        if "income_statement" in multi_year:
            self._create_statement_sheet(
                wb, "Income Statement", multi_year["income_statement"]
            )

        # 3. Balance Sheet Sheet
        if "balance_sheet" in multi_year:
            self._create_statement_sheet(
                wb, "Balance Sheet", multi_year["balance_sheet"]
            )

        # 4. Cash Flow Sheet
        if "cash_flow" in multi_year:
            self._create_statement_sheet(wb, "Cash Flow", multi_year["cash_flow"])

        # 5. Ratios Sheet
        ratios_df = report_data.get("ratios", pd.DataFrame())
        if not ratios_df.empty:
            self._create_ratios_sheet(wb, ratios_df)

        # 6. Trends Sheet
        trends_df = report_data.get("trends", pd.DataFrame())
        if not trends_df.empty:
            self._create_trends_sheet(wb, trends_df)

        # 7. Common Size Sheet
        common_size = report_data.get("common_size", {})
        if common_size:
            self._create_common_size_sheet(wb, common_size)

        # Save workbook
        wb.save(file_path_or_buffer)

    def _create_summary_sheet(
        self, wb, report: FinancialAnalysisReport, report_data: dict
    ) -> None:
        """Create summary sheet with key metrics."""
        ws = wb.create_sheet("Summary", 0)

        summary = report.get_summary_metrics()

        # Header
        ws["A1"] = "Financial Analysis Summary"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:B1")

        row = 3

        # Company Information
        ws[f"A{row}"] = "Company Information"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        ws[f"A{row}"] = "Company Name:"
        ws[f"B{row}"] = summary.get("company_name", "N/A")
        row += 1

        ws[f"A{row}"] = "Ticker:"
        ws[f"B{row}"] = summary.get("ticker", "N/A")
        row += 1

        ws[f"A{row}"] = "CIK:"
        ws[f"B{row}"] = summary.get("cik", "N/A")
        row += 1

        ws[f"A{row}"] = "Analysis Date:"
        ws[f"B{row}"] = summary.get("analysis_date", "N/A")
        row += 2

        # Key Metrics
        ws[f"A{row}"] = "Key Financial Metrics (Most Recent Period)"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        metrics = [
            ("Revenue", summary.get("revenue")),
            ("Net Income", summary.get("net_income")),
            ("Gross Profit", summary.get("gross_profit")),
            ("Operating Income", summary.get("operating_income")),
            ("Total Assets", summary.get("total_assets")),
            ("Total Liabilities", summary.get("total_liabilities")),
            ("Stockholders Equity", summary.get("equity")),
            ("Operating Cash Flow", summary.get("operating_cash_flow")),
        ]

        for label, value in metrics:
            if value is not None and not pd.isna(value):
                ws[f"A{row}"] = f"{label}:"
                if isinstance(value, (int, float)) and abs(value) > 1000:
                    ws[f"B{row}"] = f"${value:,.0f}"
                else:
                    ws[f"B{row}"] = value
                row += 1

        row += 1

        # Key Ratios
        ws[f"A{row}"] = "Key Ratios (Most Recent Period)"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        ratio_metrics = [
            ("Gross Profit Margin", summary.get("gross_margin"), "%"),
            ("Net Profit Margin", summary.get("net_margin"), "%"),
            ("Return on Assets (ROA)", summary.get("roa"), "%"),
            ("Return on Equity (ROE)", summary.get("roe"), "%"),
        ]

        for label, value, unit in ratio_metrics:
            if value is not None and not pd.isna(value):
                ws[f"A{row}"] = f"{label}:"
                ws[f"B{row}"] = f"{value:.2f}{unit}"
                row += 1

        # Format columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_statement_sheet(self, wb, sheet_name: str, df: pd.DataFrame) -> None:
        """Create a financial statement sheet."""
        ws = wb.create_sheet(sheet_name)

        # Write header
        ws["A1"] = sheet_name
        ws["A1"].font = Font(bold=True, size=14)

        # Write DataFrame
        # Reset index to include dates as first column
        df_write = df.reset_index()

        # Write column headers
        for col_idx, col_name in enumerate(df_write.columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, row_data in enumerate(df_write.values, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(value, (int, float)):
                    if abs(value) > 1000:
                        cell.value = value
                        cell.number_format = "$#,##0"
                    else:
                        cell.value = value
                        cell.number_format = "#,##0.00"
                else:
                    cell.value = value

        # Format date column
        if "end" in df_write.columns:
            date_col_idx = df_write.columns.get_loc("end") + 1
            for row in range(3, len(df_write) + 3):
                cell = ws.cell(row=row, column=date_col_idx)
                if cell.value:
                    try:
                        cell.number_format = "YYYY-MM-DD"
                    except Exception:
                        pass

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(df_write.columns, start=1):
            max_length = max(
                len(str(col_name)),
                max(
                    (
                        len(str(cell.value))
                        for cell in ws[get_column_letter(col_idx)]
                        if cell.value
                    ),
                    default=0,
                ),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

        # Freeze first row and column
        ws.freeze_panes = "B3"

    def _create_ratios_sheet(self, wb, ratios_df: pd.DataFrame) -> None:
        """Create ratios sheet."""
        ws = wb.create_sheet("Ratios")

        # Write header
        ws["A1"] = "Financial Ratios"
        ws["A1"].font = Font(bold=True, size=14)

        # Write DataFrame
        df_write = ratios_df.reset_index()

        # Write column headers
        for col_idx, col_name in enumerate(df_write.columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, row_data in enumerate(df_write.values, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(value, (int, float)) and not pd.isna(value):
                    cell.value = value
                    # Format as percentage for ratio columns
                    if "%" in str(df_write.columns[col_idx - 1]):
                        cell.number_format = "0.00%"
                    else:
                        cell.number_format = "#,##0.00"
                else:
                    cell.value = value

        # Auto-adjust column widths
        for col_idx in range(1, len(df_write.columns) + 1):
            max_length = max(
                len(str(df_write.columns[col_idx - 1])),
                max(
                    (
                        len(str(cell.value))
                        for cell in ws[get_column_letter(col_idx)]
                        if cell.value
                    ),
                    default=0,
                ),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

        ws.freeze_panes = "B3"

    def _create_trends_sheet(self, wb, trends_df: pd.DataFrame) -> None:
        """Create trends sheet with growth rates."""
        ws = wb.create_sheet("Trends")

        # Write header
        ws["A1"] = "Year-over-Year Growth Rates"
        ws["A1"].font = Font(bold=True, size=14)

        # Write DataFrame
        df_write = trends_df.reset_index()

        # Write column headers
        for col_idx, col_name in enumerate(df_write.columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, row_data in enumerate(df_write.values, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(value, (int, float)) and not pd.isna(value):
                    cell.value = value / 100  # Convert percentage to decimal for Excel
                    cell.number_format = "0.00%"
                else:
                    cell.value = value

        # Auto-adjust column widths
        for col_idx in range(1, len(df_write.columns) + 1):
            max_length = max(
                len(str(df_write.columns[col_idx - 1])),
                max(
                    (
                        len(str(cell.value))
                        for cell in ws[get_column_letter(col_idx)]
                        if cell.value
                    ),
                    default=0,
                ),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

        ws.freeze_panes = "B3"

    def _create_common_size_sheet(self, wb, common_size: dict) -> None:
        """Create common-size statements sheet."""
        ws = wb.create_sheet("Common Size")

        # Write header
        ws["A1"] = "Common-Size Financial Statements"
        ws["A1"].font = Font(bold=True, size=14)

        row = 3

        # Income Statement Common Size
        if "income_statement" in common_size:
            ws[f"A{row}"] = "Income Statement (% of Revenue)"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1

            is_df = common_size["income_statement"].reset_index()

            # Write headers
            for col_idx, col_name in enumerate(is_df.columns, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = col_name
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")

            row += 1

            # Write data
            for _, row_data in is_df.iterrows():
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        cell.value = value / 100  # Convert to decimal
                        cell.number_format = "0.00%"
                    else:
                        cell.value = value
                row += 1

            row += 2

        # Balance Sheet Common Size
        if "balance_sheet" in common_size:
            ws[f"A{row}"] = "Balance Sheet (% of Total Assets)"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1

            bs_df = common_size["balance_sheet"].reset_index()

            # Write headers
            for col_idx, col_name in enumerate(bs_df.columns, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = col_name
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")

            row += 1

            # Write data
            for _, row_data in bs_df.iterrows():
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        cell.value = value / 100
                        cell.number_format = "0.00%"
                    else:
                        cell.value = value
                row += 1

        # Auto-adjust column widths
        for col_idx in range(1, 10):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    def export_income_statement_analysis(
        self,
        income_statement_df: pd.DataFrame,
        profitability_df: pd.DataFrame,
        scale_factor: float,
        unit_label: str,
        file_path_or_buffer: Union[str, BytesIO],
    ) -> None:
        """
        Export Income Statement with profitability calculations to Excel format.

        This method creates an Excel workbook with the Income Statement formatted
        exactly as shown on the web page, including profitability calculations below.

        Args:
            income_statement_df: Income Statement DataFrame with "Metric" column and date columns
            profitability_df: Profitability calculations DataFrame with "Metric" column and date columns
            scale_factor: Scale factor to apply to values (e.g., 1e6 for millions)
            unit_label: Unit label to display (e.g., "millions")
            file_path_or_buffer: File path (string) or BytesIO buffer
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. Install it with: pip install openpyxl"
            )

        if income_statement_df.empty:
            raise ValueError("Income statement DataFrame is empty")

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income Statement Analysis"

        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center_align = Alignment(horizontal="center")
        right_align = Alignment(horizontal="right")
        left_align = Alignment(horizontal="left")

        # Final calculations that should be bold
        final_calculations = [
            "Gross Profit",
            "Operating Income",
            "Other income (expense), net",
            "Income Before Taxes",
            "Net Income",
        ]

        # Metrics that should have spacer rows after them
        spacer_metrics = [
            "Gross Profit",
            "Operating Income",
            "Other income (expense), net",
            "Income Before Taxes",
            "Net Income",
        ]

        # Bold metrics in profitability section
        bold_profitability_metrics = ["Operating Margin"]

        # Metrics that should have spacer rows after them in profitability section
        spacer_after_profitability_metrics = ["Operating Margin"]

        # Write header row
        row_num = 1
        for col_idx, col_name in enumerate(income_statement_df.columns, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            if col_idx == 1:
                # Top-left cell: "Metric" with unit below
                cell.value = f"Metric\n({unit_label})" if unit_label else "Metric"
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
            else:
                cell.value = col_name
                cell.alignment = center_align
            cell.font = header_font
            cell.fill = header_fill

        row_num += 1

        # Track row numbers for each metric (for formula references)
        metric_to_row = {}

        # Write Income Statement rows
        for idx, row_data in income_statement_df.iterrows():
            metric_name = row_data["Metric"]
            is_final_calculation = metric_name in final_calculations

            # Store row number for this metric
            metric_to_row[metric_name] = row_num

            # Write metric name
            cell = ws.cell(row=row_num, column=1)
            cell.value = metric_name
            cell.font = bold_font if is_final_calculation else normal_font
            cell.alignment = left_align

            # Write data cells
            is_eps_metric = "EPS" in metric_name

            # Determine if this is a calculated metric that should have formulas
            is_calculated_metric = metric_name in final_calculations

            for col_idx, col_name in enumerate(
                income_statement_df.columns[1:], start=2
            ):
                value = row_data[col_name]
                cell = ws.cell(row=row_num, column=col_idx)

                if pd.isna(value):
                    cell.value = "—"
                elif is_calculated_metric:
                    # Write Excel formula instead of hardcoded value
                    formula = self._get_calculation_formula(
                        metric_name, col_idx, metric_to_row, income_statement_df
                    )
                    if formula:
                        cell.value = formula
                    else:
                        # Fallback to value if formula can't be determined
                        try:
                            if is_eps_metric:
                                eps_value = float(value)
                                cell.value = eps_value
                                cell.number_format = "#,##0.00"
                            else:
                                scaled_value = float(value) / scale_factor
                                rounded_value = round(scaled_value)
                                cell.value = rounded_value
                                if rounded_value < 0:
                                    cell.number_format = "#,##0_);(#,##0)"
                                else:
                                    cell.number_format = "#,##0"
                        except (ValueError, TypeError):
                            cell.value = str(value)
                else:
                    # Non-calculated metrics: write values
                    try:
                        if is_eps_metric:
                            # EPS values are not scaled - display with 2 decimal places
                            eps_value = float(value)
                            cell.value = eps_value
                            cell.number_format = "#,##0.00"
                        else:
                            # Scale the value by the detected scale factor
                            scaled_value = float(value) / scale_factor
                            rounded_value = round(scaled_value)
                            cell.value = rounded_value
                            # Format: negative values in parentheses (Excel custom format)
                            if rounded_value < 0:
                                cell.number_format = "#,##0_);(#,##0)"
                            else:
                                cell.number_format = "#,##0"
                    except (ValueError, TypeError):
                        cell.value = str(value)

                cell.font = bold_font if is_final_calculation else normal_font
                cell.alignment = right_align

            row_num += 1

            # Add blank spacer row after key metrics
            if metric_name in spacer_metrics:
                for col_idx in range(1, len(income_statement_df.columns) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.value = ""
                row_num += 1

        # Track profitability metric rows for formula references
        # Only store FIRST occurrence (before trends section) to avoid circular references
        profitability_to_row = {}
        yoy_trend_rows = []  # Track rows in Y/Y trends section for conditional formatting

        # Add profitability calculations section
        if not profitability_df.empty:
            # Add a blank spacer row before calculated metrics
            for col_idx in range(1, len(income_statement_df.columns) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = ""
            row_num += 1

            # Track if we're in Y/Y change section for conditional formatting
            in_yoy_section = False

            # Iterate through profitability DataFrame and add rows
            for idx, row_data in profitability_df.iterrows():
                metric_name = row_data["Metric"]
                is_bold = metric_name in bold_profitability_metrics
                is_header = metric_name == "Expenses as % of Revenue"
                is_yoy_header = metric_name == "**%Change y/y Change (Trends)**"

                # Track when we enter the Y/Y change section
                if is_yoy_header:
                    in_yoy_section = True

                # Only store row number for FIRST occurrence (BEFORE trends section)
                # This prevents circular references when the same metric appears in trends
                # We explicitly check that we're NOT in the trends section yet
                if metric_name not in profitability_to_row and not in_yoy_section:
                    profitability_to_row[metric_name] = row_num

                # Write metric name (remove markdown bold markers)
                cell = ws.cell(row=row_num, column=1)
                cell.value = metric_name.replace("**", "")
                cell.font = (
                    bold_font
                    if (is_bold or is_header or is_yoy_header)
                    else normal_font
                )
                cell.alignment = left_align

                # Identify dollar amount metrics (not percentages)
                # These should NOT be conditionally formatted and should use currency formatting
                dollar_amount_metrics = {
                    "Depreciation & Amortization",
                    "CapEx",
                    "Receivables",
                    "Inventory",
                    "Payables",
                    "Change in WC",
                }
                # Identify "% of Sales" metrics from Capital & Working Capital section
                # These should NOT be conditionally formatted (they're not in trends section)
                capital_wc_percentage_metrics = {
                    "D&A % of Sales",
                    "CapEx % of Sales",
                    "Receivables % of Sales",
                    "Inventory % of Sales",
                    "Payables % of Sales",
                }
                is_dollar_amount = metric_name in dollar_amount_metrics
                is_capital_wc_percentage = metric_name in capital_wc_percentage_metrics

                # Write data cells for each date column
                for col_idx, col_name in enumerate(
                    income_statement_df.columns[1:], start=2
                ):
                    value = row_data.get(col_name)
                    cell = ws.cell(row=row_num, column=col_idx)

                    # Set format FIRST before assigning value/formula to ensure it's applied correctly
                    # This prevents any default formats from being inherited
                    if is_dollar_amount:
                        # Dollar amounts: format like income statement values (handles both positive and negative)
                        cell.number_format = "#,##0_);(#,##0)"
                    elif not is_header and not is_yoy_header:
                        # Percentages: format as percentage (only for non-header rows)
                        cell.number_format = "0.00%"

                    # Header rows should have blank cells
                    if is_header or is_yoy_header:
                        cell.value = ""
                    else:
                        # Try to write formula instead of hardcoded value
                        formula = self._get_profitability_formula(
                            metric_name,
                            col_idx,
                            col_name,
                            metric_to_row,
                            income_statement_df,
                            in_yoy_section,
                            is_yoy_header,
                            profitability_to_row,
                            row_num,
                        )

                        if formula:
                            cell.value = formula
                            cell.font = bold_font if is_bold else normal_font

                            # Track rows in Y/Y trends section for conditional formatting
                            # BUT exclude dollar amount metrics and Capital & Working Capital % metrics
                            if (
                                in_yoy_section
                                and not is_yoy_header
                                and not is_dollar_amount
                                and not is_capital_wc_percentage
                            ):
                                yoy_trend_rows.append((row_num, col_idx))
                        else:
                            # Fallback to value if formula can't be determined
                            if pd.isna(value) or value is None:
                                cell.value = "—"
                            else:
                                try:
                                    num_value = float(value)

                                    if is_dollar_amount:
                                        # Dollar amounts: scale and format like income statement
                                        # Values are already in raw dollars, so scale them
                                        scaled_value = num_value / scale_factor
                                        rounded_value = round(scaled_value)
                                        cell.value = rounded_value
                                        # Format already set above, but ensure it's correct
                                        cell.number_format = "#,##0_);(#,##0)"
                                    else:
                                        # Percentages: format as percentage
                                        cell.value = (
                                            num_value  # Store as decimal (0.50 for 50%)
                                        )
                                        # Format already set above, but ensure it's correct
                                        cell.number_format = "0.00%"

                                    # Conditional formatting for Y/Y change section ONLY
                                    # Exclude dollar amount metrics and Capital & Working Capital % metrics
                                    if (
                                        in_yoy_section
                                        and not is_yoy_header
                                        and not is_dollar_amount
                                        and not is_capital_wc_percentage
                                    ):
                                        yoy_trend_rows.append((row_num, col_idx))
                                        if num_value > 0:
                                            # Positive change - green background
                                            cell.fill = PatternFill(
                                                start_color="C6EFCE",
                                                end_color="C6EFCE",
                                                fill_type="solid",
                                            )
                                            cell.font = Font(
                                                bold=True, color="006100", size=10
                                            )
                                        elif num_value < 0:
                                            # Negative change - red background
                                            cell.fill = PatternFill(
                                                start_color="FFC7CE",
                                                end_color="FFC7CE",
                                                fill_type="solid",
                                            )
                                            cell.font = Font(
                                                bold=True, color="9C0006", size=10
                                            )
                                        else:
                                            cell.font = normal_font
                                    else:
                                        cell.font = (
                                            bold_font if is_bold else normal_font
                                        )
                                except (ValueError, TypeError):
                                    cell.value = "—"
                                    cell.font = normal_font

                    cell.alignment = right_align

                row_num += 1

                # Add spacer row after specific metrics
                if metric_name in spacer_after_profitability_metrics:
                    for col_idx in range(1, len(income_statement_df.columns) + 1):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.value = ""
                    row_num += 1

            # Apply conditional formatting to Y/Y trends section using Excel conditional formatting rules
            if yoy_trend_rows:
                self._apply_conditional_formatting(ws, yoy_trend_rows)

        # Format column widths
        ws.column_dimensions["A"].width = 35  # Metric column
        for col_idx in range(2, len(income_statement_df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15

        # Freeze first row and first column
        ws.freeze_panes = "B2"

        # Save workbook
        wb.save(file_path_or_buffer)

    def _get_calculation_formula(
        self,
        metric_name: str,
        col_idx: int,
        metric_to_row: dict,
        income_statement_df: pd.DataFrame,
    ) -> Optional[str]:
        """
        Generate Excel formula for calculated Income Statement metrics.

        Args:
            metric_name: Name of the metric (e.g., "Gross Profit")
            col_idx: Column index (Excel column number, 2-based)
            metric_to_row: Dictionary mapping metric names to row numbers
            income_statement_df: Income Statement DataFrame

        Returns:
            Excel formula string or None if formula can't be determined
        """
        from openpyxl.utils import get_column_letter

        col_letter = get_column_letter(col_idx)

        # Define calculation formulas
        formulas = {
            "Gross Profit": lambda: self._build_formula(
                col_letter, metric_to_row, ["Revenue"], ["Cost of Revenue"], "-"
            ),
            "Operating Income": lambda: self._build_formula(
                col_letter,
                metric_to_row,
                ["Gross Profit"],
                ["R&D Expenses", "SG&A Expenses"],
                "-",
            ),
            "Other income (expense), net": lambda: self._build_formula(
                col_letter,
                metric_to_row,
                ["Interest Income"],
                ["Interest Expense", "Other, net"],
                "-+",
            ),
            "Income Before Taxes": lambda: self._build_formula(
                col_letter,
                metric_to_row,
                ["Operating Income", "Other income (expense), net"],
                [],
                "+",
            ),
            "Net Income": lambda: self._build_formula(
                col_letter, metric_to_row, ["Income Before Taxes"], ["Taxes"], "-"
            ),
        }

        if metric_name in formulas:
            return formulas[metric_name]()
        return None

    def _build_formula(
        self,
        col_letter: str,
        metric_to_row: dict,
        positive_metrics: list,
        negative_metrics: list,
        operation: str,
    ) -> Optional[str]:
        """
        Build an Excel formula from metric references.

        Args:
            col_letter: Excel column letter (e.g., "B")
            metric_to_row: Dictionary mapping metric names to row numbers
            positive_metrics: List of metrics to add/subtract positively
            negative_metrics: List of metrics to subtract/add
            operation: Operation type ("-", "+", "-+")

        Returns:
            Excel formula string or None if any metric is missing
        """
        parts = []

        # Add positive metrics
        for metric in positive_metrics:
            if metric in metric_to_row:
                row_num = metric_to_row[metric]
                parts.append(f"{col_letter}{row_num}")
            else:
                return None  # Can't build formula if metric is missing

        # Add negative metrics based on operation
        if operation == "-":
            for metric in negative_metrics:
                if metric in metric_to_row:
                    row_num = metric_to_row[metric]
                    parts.append(f"-{col_letter}{row_num}")
        elif operation == "-+":
            # For "Other income (expense), net": Interest Income - Interest Expense + Other, net
            if len(negative_metrics) >= 1 and negative_metrics[0] in metric_to_row:
                row_num = metric_to_row[negative_metrics[0]]
                parts.append(f"-{col_letter}{row_num}")
            if len(negative_metrics) >= 2 and negative_metrics[1] in metric_to_row:
                row_num = metric_to_row[negative_metrics[1]]
                parts.append(f"+{col_letter}{row_num}")
        elif operation == "+":
            for metric in negative_metrics:
                if metric in metric_to_row:
                    row_num = metric_to_row[metric]
                    parts.append(f"+{col_letter}{row_num}")

        if parts:
            return "=" + "+".join(parts).replace("+-", "-")
        return None

    def _get_profitability_formula(
        self,
        metric_name: str,
        col_idx: int,
        col_name: str,
        metric_to_row: dict,
        income_statement_df: pd.DataFrame,
        in_yoy_section: bool,
        is_yoy_header: bool,
        profitability_to_row: dict,
        current_row: int,
    ) -> Optional[str]:
        """
        Generate Excel formula for profitability calculations.

        Args:
            metric_name: Name of the profitability metric
            col_idx: Column index (Excel column number, 2-based)
            col_name: Column name (date)
            metric_to_row: Dictionary mapping metric names to row numbers
            income_statement_df: Income Statement DataFrame
            in_yoy_section: Whether we're in the Y/Y change section
            is_yoy_header: Whether this is a Y/Y header row

        Returns:
            Excel formula string or None if formula can't be determined
        """
        from openpyxl.utils import get_column_letter

        col_letter = get_column_letter(col_idx)

        # Get column index in DataFrame for finding previous period
        date_columns = list(income_statement_df.columns[1:])
        try:
            current_col_idx = date_columns.index(col_name)
        except ValueError:
            return None

        # Y/Y Growth formulas
        if "Y/Y % Change" in metric_name and not in_yoy_section:
            # Find the metric being compared (e.g., "Revenue Y/Y % Change" -> "Revenue")
            base_metric = metric_name.replace(" Y/Y % Change", "")
            if base_metric in metric_to_row:
                row_num = metric_to_row[base_metric]
                # Compare to previous period (next column)
                if current_col_idx + 1 < len(date_columns):
                    prev_col_letter = get_column_letter(col_idx + 1)
                    # Formula: (Current - Previous) / Previous
                    return f"=IF({prev_col_letter}{row_num}<>0,({col_letter}{row_num}-{prev_col_letter}{row_num})/{prev_col_letter}{row_num},0)"
            return None

        # Y/Y Change Trends (difference between consecutive growth rates)
        if in_yoy_section and not is_yoy_header:
            # Handle "Change of Y/Y % Change" metrics (e.g., "Outstanding Shares Basic Change of Y/Y % Change")
            if "Change of Y/Y % Change" in metric_name:
                # Extract base metric name (e.g., "Outstanding Shares Basic" from "Outstanding Shares Basic Change of Y/Y % Change")
                base_metric_name = metric_name.replace(" Change of Y/Y % Change", "")
                yoy_metric_name = f"{base_metric_name} Y/Y % Change"

                # Reference the Y/Y % Change row from the first section
                if yoy_metric_name in profitability_to_row:
                    yoy_row = profitability_to_row[yoy_metric_name]
                    # CRITICAL: Make sure we're referencing a row from BEFORE the trends section
                    if (
                        yoy_row != current_row
                        and yoy_row < current_row
                        and current_col_idx + 1 < len(date_columns)
                    ):
                        prev_col_letter = get_column_letter(col_idx + 1)
                        # Formula: Current Y/Y % Change - Previous Y/Y % Change
                        return f"={col_letter}{yoy_row}-{prev_col_letter}{yoy_row}"

            # Handle "Revenue Y/Y % Change" in trends section (references the first "Revenue Y/Y % Change" row)
            elif (
                metric_name == "Revenue Y/Y % Change"
                and "Revenue Y/Y % Change" in profitability_to_row
            ):
                yoy_row = profitability_to_row["Revenue Y/Y % Change"]
                # CRITICAL: Make sure we're referencing a row from BEFORE the trends section
                if (
                    yoy_row != current_row
                    and yoy_row < current_row
                    and current_col_idx + 1 < len(date_columns)
                ):
                    prev_col_letter = get_column_letter(col_idx + 1)
                    # Formula: Current Y/Y % Change - Previous Y/Y % Change
                    # Note: This references the first occurrence of "Revenue Y/Y % Change" from above
                    return f"={col_letter}{yoy_row}-{prev_col_letter}{yoy_row}"

            # For percentage metrics in trends section (e.g., "Gross Margin", "Operating Margin")
            # These are the Y/Y change in percentage points
            # IMPORTANT: These must reference the ORIGINAL % of Revenue rows, not the trends rows
            else:
                # Check if this percentage metric exists in profitability rows (from the % of Revenue section)
                if metric_name in profitability_to_row:
                    pct_row = profitability_to_row[metric_name]
                    # CRITICAL: Make absolutely sure we're not referencing the current row
                    # The stored row should be from BEFORE the trends section
                    if (
                        pct_row != current_row
                        and pct_row < current_row
                        and current_col_idx + 1 < len(date_columns)
                    ):
                        prev_col_letter = get_column_letter(col_idx + 1)
                        # Formula: Current % - Previous % (from the original % of Revenue section)
                        # This calculates the change in percentage points between periods
                        return f"={col_letter}{pct_row}-{prev_col_letter}{pct_row}"

            return None

        # Percentage of Revenue formulas
        if "Expenses as % of Revenue" not in metric_name and "%" not in metric_name:
            # Map metric names to their numerator/denominator
            percentage_formulas = {
                "Gross Margin": ("Gross Profit", "Revenue"),
                "Research and development": ("R&D Expenses", "Revenue"),
                "Sales, general and administrative": ("SG&A Expenses", "Revenue"),
                "Restructuring and other charges": (
                    "Restructuring and other charges",
                    "Revenue",
                ),
                "Acquisition termination cost": (
                    "Acquisition termination cost",
                    "Revenue",
                ),
                "Interest income": ("Interest Income", "Revenue"),
                "Interest expense": ("Interest Expense", "Revenue"),
                "Other, net": ("Other, net", "Revenue"),
                "Operating Margin": ("Operating Income", "Revenue"),
                "Tax rate": ("Taxes", "Income Before Taxes"),
            }

            for key, (numerator, denominator) in percentage_formulas.items():
                if key.lower() in metric_name.lower():
                    if numerator in metric_to_row and denominator in metric_to_row:
                        num_row = metric_to_row[numerator]
                        denom_row = metric_to_row[denominator]
                        # Formula: Numerator / Denominator
                        return f"=IF({col_letter}{denom_row}<>0,{col_letter}{num_row}/{col_letter}{denom_row},0)"
            return None

        # Handle percentage metrics from new Capital & Working Capital section
        # These are calculated as: metric / Revenue
        percentage_metrics = {
            "D&A % of Sales": ("Depreciation & Amortization", "Revenue"),
            "CapEx % of Sales": ("CapEx", "Revenue"),
            "Receivables % of Sales": ("Receivables", "Revenue"),
            "Inventory % of Sales": ("Inventory", "Revenue"),
            "Payables % of Sales": ("Payables", "Revenue"),
        }

        for pct_metric, (numerator, denominator) in percentage_metrics.items():
            if metric_name == pct_metric:
                # Find numerator row in profitability_to_row (from Capital & Working Capital section)
                # Find denominator row in metric_to_row (from Income Statement)
                if numerator in profitability_to_row and denominator in metric_to_row:
                    num_row = profitability_to_row[numerator]
                    denom_row = metric_to_row[denominator]
                    # Formula: Numerator / Denominator
                    return f"=IF({col_letter}{denom_row}<>0,{col_letter}{num_row}/{col_letter}{denom_row},0)"
                return None

        # Handle Change in WC formula
        # Change in WC = (Receivables + Inventory - Payables) current - (Receivables + Inventory - Payables) previous
        if metric_name == "Change in WC":
            receivables_row = profitability_to_row.get("Receivables")
            inventory_row = profitability_to_row.get("Inventory")
            payables_row = profitability_to_row.get("Payables")

            if receivables_row and inventory_row and payables_row:
                # Check if there's a previous period
                if current_col_idx + 1 < len(date_columns):
                    prev_col_letter = get_column_letter(col_idx + 1)
                    # Formula: (Receivables + Inventory - Payables) current - (Receivables + Inventory - Payables) previous
                    current_wc = f"({col_letter}{receivables_row}+{col_letter}{inventory_row}-{col_letter}{payables_row})"
                    prev_wc = f"({prev_col_letter}{receivables_row}+{prev_col_letter}{inventory_row}-{prev_col_letter}{payables_row})"
                    return f"={current_wc}-{prev_wc}"
                else:
                    # No previous period, return empty
                    return None

        # Handle Y/Y change for new percentage metrics in trends section
        if in_yoy_section and not is_yoy_header:
            for pct_metric in percentage_metrics.keys():
                if metric_name == pct_metric:
                    # Reference the original percentage row from Capital & Working Capital section
                    if pct_metric in profitability_to_row:
                        pct_row = profitability_to_row[pct_metric]
                        # CRITICAL: Make sure we're referencing a row from BEFORE the trends section
                        if (
                            pct_row != current_row
                            and pct_row < current_row
                            and current_col_idx + 1 < len(date_columns)
                        ):
                            prev_col_letter = get_column_letter(col_idx + 1)
                            # Formula: Current % - Previous % (from the original % of Sales section)
                            return f"={col_letter}{pct_row}-{prev_col_letter}{pct_row}"
                    return None

        return None

    def _apply_conditional_formatting(self, ws, yoy_trend_rows):
        """
        Apply conditional formatting to Y/Y trends section cells.

        Uses Excel's conditional formatting rules to color cells:
        - Green background for positive values
        - Red background for negative values

        Args:
            ws: Worksheet object
            yoy_trend_rows: List of (row_num, col_idx) tuples for cells to format
        """
        try:
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import PatternFill, Font

            if not yoy_trend_rows:
                return

            # Group cells by column for efficient formatting
            from collections import defaultdict

            cells_by_col = defaultdict(list)

            for row_num, col_idx in yoy_trend_rows:
                cells_by_col[col_idx].append(row_num)

            # Apply conditional formatting to each column
            for col_idx, row_nums in cells_by_col.items():
                from openpyxl.utils import get_column_letter

                col_letter = get_column_letter(col_idx)

                # Create range for this column
                min_row = min(row_nums)
                max_row = max(row_nums)
                cell_range = f"{col_letter}{min_row}:{col_letter}{max_row}"

                # Green fill for positive values
                green_fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                green_font = Font(bold=True, color="006100", size=10)
                positive_rule = CellIsRule(
                    operator="greaterThan",
                    formula=["0"],
                    fill=green_fill,
                    font=green_font,
                )
                ws.conditional_formatting.add(cell_range, positive_rule)

                # Red fill for negative values
                red_fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
                red_font = Font(bold=True, color="9C0006", size=10)
                negative_rule = CellIsRule(
                    operator="lessThan", formula=["0"], fill=red_fill, font=red_font
                )
                ws.conditional_formatting.add(cell_range, negative_rule)
        except ImportError:
            # If conditional formatting is not available, skip it
            pass
